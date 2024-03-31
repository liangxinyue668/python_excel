import openpyxl
import openpyxl.styles

excel_1 = openpyxl.load_workbook('./作业.xlsx')
sheet_1 = excel_1.get_sheet_by_name('Sheet1')
sheet_2 = excel_1.get_sheet_by_name('Sheet2')
sheet_3 = excel_1['Sheet3']

dic={}
for i in range(1, sheet_1.max_column+1):
    dic[sheet_1.cell(1, i).value] = i


titles = [
    'A',
	'B',
	'C',
	'D',
	'E']

for i in range(len(titles)):
    sheet_3.cell(1, i+1, titles[i])



data = []

for i in range(2, sheet_1.max_row+2):    
    row_1 = []
    row_2 = []
    row_3 = []
    for j in range(len(titles)):
        row_3.append(sheet_1.cell(i, j+1).value)
        row_1.append(sheet_2.cell(i*2-2, 5-j).value)
        row_2.append(sheet_2.cell(i*2-1, 5-j).value)
    data.append(row_3)
    data.append(row_1)
    data.append(row_2)

for i in range(len(data)):
    for j in range(len(data[i])):
        print(i, j)
        sheet_3.cell(i+2, j+1, data[i][j])

for cell in sheet_3[1]:
    cell.font = openpyxl.styles.Font(bold = True)

for row in sheet_3.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

excel_1.save('./作业.xlsx')

excel_1.close()



