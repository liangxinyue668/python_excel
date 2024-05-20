import openpyxl

excel_1 = openpyxl.load_workbook('./文件1.xlsx')
excel_2 = openpyxl.load_workbook('./文件2.xlsx')
sheet_1 = excel_1['Sheet1']
sheet_2 = excel_2['Sheet1']
sheet_3 = excel_2['Sheet2']

data_1 = []

#文件2第一个表格的数据需要移动到第二个表格的值
for i in range(1, 6):
    row_1 = []
    for j in range(1, 8):
        row_1.append(sheet_2.cell(i, j).value)
    data_1.append(row_1)

for i in range(len(data_1)):
    for j in range(len(data_1[i])):
        sheet_3.cell(i+8, j+1, data_1[i][j])

data_2 = []

#文件1数据复制到文件2
for i in range(1, 6):
    row_2 = [] 
    for j in range(1, 8):
        row_2.append(sheet_1.cell(i, j).value)
    data_2.append(row_2)

#将数据保存到文件2的sheet2
for i in range(len(data_2)):
    for j in range(len(data_2[i])):
        sheet_3.cell(i+1, j+1, data_2[i][j])

excel_2.save('./文件2.xlsx')

excel_2.close()


