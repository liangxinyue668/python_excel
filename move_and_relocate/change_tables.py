import openpyxl
import openpyxl.styles

excel_1 = openpyxl.load_workbook('./文件2.xlsx')
sheet_1 = excel_1['Sheet1']
sheet_2 = excel_1['Sheet2']


def change_table(source_sheet, target_sheet, start_source_row, start_source_col, start_target_row, start_target_col, row_num, col_num):
    
    data = []
    
    # 从工作表读取数据
    for i in range(start_source_row, start_source_row + row_num):
        row = []
        for j in range(start_source_col, start_source_col + col_num):
            row.append(source_sheet.cell(i,j).value)
        data.append(row)
    
    # 将数据写入目标工作表
        for i in range(len(data)):
            for j in range(len(data[i])):
                target_sheet.cell(i+start_target_row, j+start_target_col, data[i][j])
    
    # 数据居中，表头加粗
    for cell in target_sheet[start_target_row]:
        cell.font = openpyxl.styles.Font(bold = True)

    for row in target_sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')

excel_2 = openpyxl.Workbook()
sheet_3 = excel_2.active
sheet_3.title = 'Sheet1'
sheet_4 = excel_2.create_sheet('Sheet2')

# 将文件2的sheet1的第一个表读取写入文件3的sheet2的第二个表格
change_table(sheet_1, sheet_4, 1, 1, 1, 1, 5, 7)

# 将文件2的sheet1的第二个表读取写入文件3的sheet2第二个表格
change_table(sheet_1, sheet_4, 8, 1, 8, 1, 5, 7)

# 将文件2的sheet2的第一个表读取写入文件3的sheet1的第一个表格
change_table(sheet_2, sheet_3, 1, 1, 1, 1, 5, 7)

# 将文件2的sheet2的第二个表读取写入文件3的sheet1的第二个表格
change_table(sheet_2, sheet_3, 8, 1, 8, 1, 5, 7)



excel_2.save('./文件3.xlsx')


