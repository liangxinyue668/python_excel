import openpyxl
import excel_tool

week_report = openpyxl.load_workbook('D:/数据/周报/周报/亲子周报2024-05-20-2024-05-26.xlsx')
week_report_family_data = week_report['亲子周度数据']
week_report_family_inventory = week_report['亲子库存']
week_report_family_country = week_report['亲子站点数据']
week_report_shoes_data = week_report['童鞋周度数据']
week_report_shoes_inventory = week_report['童鞋库存']
week_report_shoes_country = week_report['童鞋站点数据']
week_report_acc_data = week_report['配饰周度数据']
week_report_acc_inventory = week_report['配饰库存']
week_report_acc_country = week_report['配饰站点数据']

week_pivot_table = openpyxl.load_workbook(r'D:\数据\周报\数据\亲子周报透视数据2024-05-20_2024-05-26.xlsx')
week_pivot_table_family_sales = week_pivot_table['亲子销售透视']
week_pivot_table_family_inventory = week_pivot_table['亲子库存透视']
week_pivot_table_shoes_sales = week_pivot_table['童鞋销售']
week_pivot_table_shoes_inventory = week_pivot_table['童鞋库存']
week_pivot_table_acc_sales = week_pivot_table['配饰销售']
week_pivot_table_acc_inventory = week_pivot_table['配饰库存']

week_country = openpyxl.load_workbook(r'D:\数据\周报\数据\站点商品数据2024-05-20_2024-05-26.xlsx')
week_country_family = week_country['亲子']
week_country_shoes = week_country['童鞋']
week_country_acc = week_country['配饰']

# 亲子周报
# 新老品爆旺数据
excel_tool.copy_data(week_report_family_data, week_report_family_data, 51, 2, 58, 2, 4, 15)

excel_tool.copy_data(week_pivot_table_family_sales, week_report_family_data, 38, 2, 51, 2, 4, 15)

# 新老品销量层级
excel_tool.copy_data(week_report_family_data, week_report_family_data, 75, 3, 97, 3, 17, 15)

excel_tool.copy_data(week_pivot_table_family_sales, week_report_family_data, 49, 3, 75, 3, 17, 15)

#三级类目销售
excel_tool.copy_data(week_report_family_data, week_report_family_data, 143, 3, 168, 3, 21, 15)

excel_tool.copy_data(week_pivot_table_family_sales, week_report_family_data, 77, 3, 143, 3, 21, 15)

#节日
excel_tool.copy_data(week_report_family_data, week_report_family_data, 220, 2, 233, 2, 7, 15)

excel_tool.copy_data(week_pivot_table_family_sales, week_report_family_data, 106, 2, 220, 2, 7, 15)

#生命周期销售
excel_tool.copy_data(week_report_family_data, week_report_family_data, 262, 2, 270, 2, 5, 15)

excel_tool.copy_data(week_pivot_table_family_sales, week_report_family_data, 124, 2, 262, 2, 5, 15)

#亲子库存
#三级类目库存
excel_tool.copy_data(week_report_family_inventory, week_report_family_inventory, 4, 2, 18, 2, 9, 9)

excel_tool.copy_data(week_pivot_table_family_inventory, week_report_family_inventory, 6, 2, 4, 2, 9, 9)

#季节库存
excel_tool.copy_data(week_report_family_inventory, week_report_family_inventory, 34, 2, 43, 2, 6, 9)

excel_tool.copy_data(week_pivot_table_family_inventory, week_report_family_inventory, 24, 2, 34, 2, 6, 9)

#节日品库存
excel_tool.copy_data(week_report_family_inventory, week_report_family_inventory, 53, 2, 66, 2, 7, 9)

excel_tool.copy_data(week_pivot_table_family_inventory, week_report_family_inventory, 38, 2, 53, 2, 7, 9)

#新老品库存
excel_tool.copy_data(week_report_family_inventory, week_report_family_inventory, 81, 2, 89, 2, 5, 9)

excel_tool.copy_data(week_pivot_table_family_inventory, week_report_family_inventory, 55, 2, 81, 2, 5, 9)

#生命周期库存
excel_tool.copy_data(week_report_family_inventory, week_report_family_inventory, 99, 2, 108, 2, 6, 9)

excel_tool.copy_data(week_pivot_table_family_inventory, week_report_family_inventory, 67, 2, 99, 2, 6, 9)

#亲子站点数据
excel_tool.copy_data(week_report_family_country, week_report_family_country, 5, 15, 40, 15, 31, 14)
excel_tool.copy_data(week_report_family_country, week_report_family_country, 5, 8, 40, 8, 31, 4)
excel_tool.copy_data(week_country_family, week_report_family_country, 5, 3, 5, 15, 31, 14)
excel_tool.copy_data(week_country_family, week_report_family_country, 127, 3, 5, 8, 31, 4)

week_report.save('D:\数据\周报\周报\亲子周报2024-05-20-2024-05-26.xlsx')

week_report.close()


