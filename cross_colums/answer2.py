import openpyxl
import openpyxl.styles

excel_1 = openpyxl.load_workbook('./作业2.xlsx')
sheet_1 = excel_1['Sheet1']
sheet_2 = excel_1['Sheet2']
sheet_4 = excel_1.create_sheet("Sheet4")

data = []

for i in range(1, 5):
    col_1 = []
    col_2 = []
    col_3 = []
    for j in range(1, 14):
        col_1.append(sheet_1.cell(j, i).value)
        col_2.append(sheet_2.cell(j, i*2-1).value)
        col_3.append(sheet_2.cell(j, i*2).value)
    data.append(col_1)
    data.append(col_2)
    data.append(col_3)

for i in range(len(data)):
    for j in range(len(data[i])):
        sheet_4.cell(j+1,i+1,data[i][j])

for cell in sheet_4[1]:
    cell.font = openpyxl.styles.Font(bold = True)
 
for row in sheet_4.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    


excel_1.save('./作业2.xlsx')

excel_1.close()
    