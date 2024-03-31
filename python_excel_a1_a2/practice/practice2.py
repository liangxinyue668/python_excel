
# from openpyxl import Workbook, load_workbook
import openpyxl

excel_1=openpyxl.load_workbook('./practice/运营表-test 去重版.xlsx')
sheet_1=excel_1.get_sheet_by_name('已去重')

excel_2 = openpyxl.Workbook()
sheet_2 = excel_2.create_sheet('aa')

dic={}
for i in range(1, sheet_1.max_column+1):
    dic[sheet_1.cell(1, i).value] = i

titles = [
    '一级类目辅助列',
    '所属类目辅助列',
    '商品id',
    '库存',
    '库龄大于180天的库存数量',
    '库龄0~30天的库存数量',
    '库龄30~60天的库存数量',
    '库龄60~90天的库存数量',
    '库龄90~120天的库存数量',
    '库龄120~150天的库存数量',
    '库龄150~180天的库存数量',
]

for i in range(1,len(titles)):
    sheet_2.cell(1, i, titles[i])


old_data = []

for i in range(2, sheet_1.max_row+1):
    row = []
    for j in range(len(titles)):
        row.append(sheet_1.cell(i, dic[titles[j]]).value)
    old_data.append(row)

    row = []
    row.append(sheet_1.cell(i, dic['一级类目辅助列']).value)
    row.append('三方-TEMU')
    row.append(sheet_1.cell(i, dic['商品id']).value)
    row.append(sheet_1.cell(i, dic['temu_逻辑售卖数']).value)
    old_data.append(row)

# print(old_data)

for i in range(len(old_data)):
    for j in range(len(old_data[i])):
        print(i, j)
        sheet_2.cell(i+2, j+1, old_data[i][j])


excel_2.save('./运营分析报表（含temu）去重版.xlsx')
# excel_2.save('./运营分析报表（含temu）.xlsx')
