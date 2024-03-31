import xlrd
# from openpyxl import Workbook, load_workbook
import openpyxl

# wb = openpyxl.load_workbook('filename.xlsx')
# wb = openpyxl.Workbook(write_only=True)


excel_1=xlrd.open_workbook('./practice/运营表-test.xlsx')
sheet_1=excel_1.sheet_by_name('Sheet1')

excel_2 = openpyxl.Workbook()
sheet_2 = excel_2.create_sheet('aa')

dic={}
for i in range(0, sheet_1.ncols):
    dic[sheet_1.cell_value(0, i)] = i

titles = [
    '商品id',
    '所属类目辅助列',
    '一级类目辅助列',
    '库存',
    '库龄大于180天的库存数量',
    '库龄0~30天的库存数量',
    '库龄30~60天的库存数量',
    '库龄60~90天的库存数量',
    '库龄90~120天的库存数量',
    '库龄120~150天的库存数量',
    '库龄150~180天的库存数量',
]

for i in range(len(titles)):
    sheet_2.cell(1, i+1, titles[i])


old_data = []

for i in range(1, sheet_1.nrows):
    row = []
    for j in range(len(titles)):
        row.append(sheet_1.cell_value(i, dic[titles[j]]))
    old_data.append(row)

    row = []
    row.append(sheet_1.cell_value(i, dic['商品id']))
    row.append('三方-TEMU')
    row.append(sheet_1.cell_value(i, dic['一级类目辅助列']))
    row.append(sheet_1.cell_value(i, dic['temu_逻辑售卖数']))
    old_data.append(row)

# print(old_data)

for i in range(len(old_data)):
    for j in range(len(old_data[i])):
        print(i, j)
        sheet_2.cell(i+2, j+1, old_data[i][j])


excel_2.save('./运营分析报表（含temu）.xlsx')
# excel_2.save('./运营分析报表（含temu）.xlsx')
